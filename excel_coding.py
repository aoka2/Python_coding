from openpyxl import Workbook #エクセル編集用のimport
from openpyxl.styles import Font #フォント用のimport
from openpyxl import load_workbook #エクセルファイルを読み込む用のimport
from openpyxl.styles import Side, Border #枠線を作るためのimport
from openpyxl.chart import BarChart, Reference 


wb = load_workbook(filename='testSheet.xlsx')
sheet_name = wb.sheetnames[0]
ws = wb[sheet_name]
#ws = wb.active

x = Reference(ws, min_col = 2, min_row = 2, max_col = 2, max_row = 13)

data = Reference(ws, min_col = 3, min_row =2, max_col = 3, max_row = 13)

chart = BarChart() #ここをLineChartにすれば棒グラフ pieChartにすれば円グラフになる
chart.add_data(data)
chart.set_categories(x)
ws.add_chart(chart,'E2')
#sheet_name = wb.sheetnames[0]
#ws1 = wb[sheet_name]
#ws1.merge_cells('A5:B7')
#ws1['A5'] = 'サプー' #結合したものの一番左上を指定して値を入れる

#s = Side(style = 'thin')
#b = Border(left=s, right=s, top=s, bottom=s)

#cell = ws1['B2']
#cell.border = b

#ws1.insert_cols(1,3) #データ列の挿入
#ws1.delete_cols(2,5) #データ列の消去
#ws1['A4'] = 'aiueo'
#x = ws1['A4'].value
#ws2 = wb.create_sheet('new sheet')
#ws2['A1'] = x
#ws = wb.active #これでシートのオブジェクトが作れる
#ws['A1'] = 10

#cel = ws['A1']
#cel.font = Font(size = 12, bold = True, color='FF0000')
wb.save('output.xlsx')
